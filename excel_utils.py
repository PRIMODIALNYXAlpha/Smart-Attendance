import os
from openpyxl import Workbook, load_workbook

FILE_NAME = "attendance.xlsx"

def create_excel():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Attendance"
        sheet.append(["Name", "Email", "Branch", "Date", "Time", "Status"])
        wb.save(FILE_NAME)

def mark_attendance_excel(name, email, branch, date, time):
    create_excel()

    wb = load_workbook(FILE_NAME)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == name and str(row[3]) == str(date):
            return

    sheet.append([name, email, branch, date, time, "Present"])
    wb.save(FILE_NAME)